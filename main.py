import imaplib
import email
from email.header import decode_header
import base64
import re
from datetime import datetime


from bs4 import BeautifulSoup
import xlrd
import xlrd3
import openpyxl
import xlwt

import config
import for_api
import filter


# Настройка imaplib
mail_pass = config.password
username = config.address
imap_server = "imap.mail.ru"
imap = imaplib.IMAP4_SSL(imap_server)
imap.login(username, mail_pass)

# imap.select("INBOX")
imap.select('user')
# p = imap.search(None, 'ALL')
# r = imap.uid('search', "UNSEEN", "ALL")
typ, data = imap.uid('search', "UNSEEN", "ALL")

def start():
    list_all_north = []
    list_all_south = []
    list_all_west = []
    list_all_east = []
    list_all_hz = []
    for num in data[0].split():
        res, msg = imap.uid('fetch', num, '(RFC822)')
        # print(num)

        mail = email.message_from_bytes(msg[0][1])

        if mail.is_multipart():
            for part in mail.walk():
                content_type = part.get_content_type()
                filename = part.get_filename()
                # Найдем файлы с названием(строка) и с 16 символом с, это econtracts
                if type(filename) == str and filename[16] == 'c':
                    print(filename)
                    print(type(filename))
                    with open(f'files/{filename}', 'wb') as new_file:
                        new_file.write(part.get_payload(decode=True))

                    # Версия для xlrd
                    # workbook = xlrd.open_workbook(f'files/{filename}')
                    # worksheet = workbook.sheet_by_index(0)

                    # Версия для openpyxl
                    # workbook_all = openpyxl.load_workbook(f'files/{filename}')
                    # sheet_all = workbook_all.active
                    # for row in sheet_all.iter_rows(values_only=True):

                    # Версия для xlrd3
                    wb = xlrd3.open_workbook(f'files/{filename}')
                    sheet = wb.sheet_by_index(0)
                    # Старт со второй строчки
                    for row in range(1, sheet.nrows):
                        list_one = []
                        # t_o = ""
                        # read_one(row, t_o)
                        # print(row)
                        # 1 Бренд, получим с помощью API
                        try: list_one.append(for_api.search_brand(int(sheet.cell_value(row, 1))))
                        except ValueError: list_one.append(" ")

                        # 2 Дата
                        list_one.append(sheet.cell_value(row, 0))

                        # 3 Лицевой счет
                        try: list_one.append(int(sheet.cell_value(row, 1)))
                        except ValueError: list_one.append(sheet.cell_value(row, 1))

                        # 4 Номер заявки
                        try: list_one.append(int(sheet.cell_value(row, 2)))
                        except ValueError: list_one.append(sheet.cell_value(row, 2))

                        # 5 Улица
                        street = filter.filter_street(sheet.cell_value(row, 3))
                        list_one.append(street)

                        # 6 Дом
                        try: list_one.append(int(sheet.cell_value(row, 4)))
                        except ValueError: list_one.append(sheet.cell_value(row, 4))

                        # 7 Квартира
                        try: list_one.append(int(sheet.cell_value(row, 5)))
                        except ValueError: list_one.append(sheet.cell_value(row, 5))

                        # 8 Мастер
                        master = sheet.cell_value(row, 6)
                        if master in filter.filter_master_no_to:
                            continue
                        else:
                            list_one.append(sheet.cell_value(row, 6))

                        # Проверим столбец оборудования до выставления типа договора
                        router = filter.filter_router(sheet.cell_value(row, 8))

                        # 9 Тип договора
                        if router == "Услуга":
                            list_one.append("Услуга")
                        else:
                            list_one.append(sheet.cell_value(row, 7))

                        # 10 Оборудование
                        if router == "Услуга":
                            list_one.append("")
                        else:
                            list_one.append(router)

                        # Итог
                        if master in filter.filter_master_north:
                            list_all_north.append(list_one)
                        elif master in filter.filter_master_south:
                            list_all_south.append(list_one)
                        elif master in filter.filter_master_west:
                            list_all_west.append(list_one)
                        elif master in filter.filter_master_east:
                            list_all_east.append(list_one)
                        else:
                            list_all_hz.append(list_one)

    # print(list_all)
    save_to_exel(list_all_north, "north")
    save_to_exel(list_all_south, "south")
    save_to_exel(list_all_west, "west")
    save_to_exel(list_all_east, "east")
    save_to_exel(list_all_hz, "hz")


def save_to_exel(list_to_exel, to):
    wb = xlwt.Workbook()
    ws = wb.add_sheet("Электронные акты")

    list_to_exel.reverse()

    for n, v in enumerate(list_to_exel):
        ws.write(n+1, 0, v[0])  # Бренд
        ws.write(n+1, 1, v[1])  # Дата
        ws.write(n+1, 2, v[2])  # Лицевой счет
        ws.write(n+1, 3, v[3])  # Номер заявки
        ws.write(n+1, 4, v[4])  # Улица
        ws.write(n+1, 5, v[5])  # Дом
        ws.write(n+1, 6, v[6])  # Квартира
        ws.write(n+1, 7, v[7])  # Мастер
        ws.write(n+1, 9, v[8])  # Тип договора
        ws.write(n+1, 10, v[9])  # Оборудование


    date_now = datetime.now()
    date_now_year = date_now.strftime("%d.%m.%Y %H:%M")


    wb.save(f'result/{to}_{date_now_year}.xlsx')

if __name__ == '__main__':
    start()
